import json
import openpyxl
import requests
import configparser
from utilities.configurations import *
from utilities.payload import *

'''
Adding multiple students and verifying them thru the get response
'''


def test_add_multiple_students_excel():
    # url = "https://thetestingworldapi.com/api/studentsDetails"
    url = getConfig()['API']['endpoint']+'api/studentsDetails'

    file = open("/Users/tolik/PycharmProjects/API_Testing/TASK_API/request_json.json", 'r')

    # Excel Code
    wk = openpyxl.load_workbook("/Users/tolik/PycharmProjects/Project_API_Example/TestCases/testData.xlsx")
    sh = wk['Sheet1']
    rows = sh.max_row
    json_request = json.loads(file.read())

    for i in range(2, rows + 1):
        cell_first_name = sh.cell(row=i, column=1)
        cell_mid_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_dob = sh.cell(row=i, column=4)

        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_mid_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_dob.value

        response = requests.post(url, json_request,)

        print(response.text)
        print(response.status_code)
        assert response.status_code == 201

    file.close()


def test_get_all_students_list():
    # url = "https://thetestingworldapi.com/api/studentsDetails"
    url = getConfig()['API']['endpoint']+'api/studentsDetails'

    response = requests.get(url)
    json_response = response.json()

    print(f"json_response type: {type(json_response)}")     # class list

    print("\n***** Response Status Code *****")
    print(f"Response Status Code: {response.status_code}")

    print("\n***** For Easy Viewing Readable Response *****")
    data = json.dumps(json_response, indent=2, sort_keys=True)
    print(type(data))       # type string

    print(data)

